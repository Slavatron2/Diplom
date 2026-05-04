import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
import webbrowser
import threading
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse, parse_qs
import time
import traceback
import mysql.connector
from mysql.connector import Error
import os
import subprocess
import sys

# ==================== КОНФИГУРАЦИЯ ПОДКЛЮЧЕНИЯ ====================
# ⚠️ В ЭТОЙ СТРОКЕ ЗАМЕНИ НА СВОЙ ПАРОЛЬ ОТ MYSQL ⚠️
MYSQL_PASSWORD = "1234"  # <-- ПРИМЕР: "root" или "123456"

# Настройки для базы данных пользователей
USERS_DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': MYSQL_PASSWORD,
    'database': 'users_db'
}

# Настройки для базы данных истории поиска
HISTORY_DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': MYSQL_PASSWORD,
    'database': 'search_history_db'
}


# ==================== ФУНКЦИИ ДЛЯ БАЗЫ ПОЛЬЗОВАТЕЛЕЙ ====================
def get_users_connection():
    """Подключение к базе данных пользователей"""
    try:
        conn = mysql.connector.connect(**USERS_DB_CONFIG)
        return conn
    except Error as e:
        print(f"Ошибка подключения к БД пользователей: {e}")
        return None


def user_exists(login):
    """Проверяет, существует ли пользователь"""
    conn = get_users_connection()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE login = %s", (login,))
        result = cursor.fetchone() is not None
        cursor.close()
        conn.close()
        return result
    except Error as e:
        print(f"Ошибка проверки пользователя: {e}")
        return False


def register_user(login, password):
    """Регистрация нового пользователя"""
    if user_exists(login):
        return False

    conn = get_users_connection()
    if not conn:
        return False

    try:
        cursor = conn.cursor()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute(
            "INSERT INTO users (login, password_hash) VALUES (%s, %s)",
            (login, password_hash)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Error as e:
        print(f"Ошибка регистрации: {e}")
        return False


def check_login(login, password):
    """Проверка логина и пароля"""
    conn = get_users_connection()
    if not conn:
        return False

    try:
        cursor = conn.cursor()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute(
            "SELECT id FROM users WHERE login = %s AND password_hash = %s",
            (login, password_hash)
        )
        result = cursor.fetchone() is not None
        cursor.close()
        conn.close()
        return result
    except Error as e:
        print(f"Ошибка проверки логина: {e}")
        return False


# ==================== ФУНКЦИИ ДЛЯ БАЗЫ ИСТОРИИ ПОИСКА ====================
def get_history_connection():
    """Подключение к базе данных истории поиска"""
    try:
        conn = mysql.connector.connect(**HISTORY_DB_CONFIG)
        return conn
    except Error as e:
        print(f"Ошибка подключения к БД истории: {e}")
        return None


def insert_search_query(user_login, query_text):
    """Сохраняет поисковый запрос в базу"""
    conn = get_history_connection()
    if not conn:
        return None

    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO search_queries (user_login, query_text) VALUES (%s, %s)",
            (user_login, query_text)
        )
        conn.commit()
        query_id = cursor.lastrowid
        cursor.close()
        conn.close()
        return query_id
    except Error as e:
        print(f"Ошибка сохранения запроса: {e}")
        return None


def insert_search_results(query_id, results):
    """Сохраняет результаты поиска в базу"""
    conn = get_history_connection()
    if not conn:
        return False

    try:
        cursor = conn.cursor()
        for rank, item in enumerate(results, start=1):
            cursor.execute('''
                INSERT INTO search_results (query_id, title, description, url, rank_position)
                VALUES (%s, %s, %s, %s, %s)
            ''', (query_id, item["Заголовок"], item["Описание"], item["Ссылка"], rank))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Error as e:
        print(f"Ошибка сохранения результатов: {e}")
        return False


def get_user_history(user_login):
    """Получает историю запросов пользователя"""
    conn = get_history_connection()
    if not conn:
        return []

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sq.id, sq.query_text, sq.search_date, COUNT(sr.id) as results_count
            FROM search_queries sq
            LEFT JOIN search_results sr ON sq.id = sr.query_id
            WHERE sq.user_login = %s
            GROUP BY sq.id, sq.query_text, sq.search_date
            ORDER BY sq.search_date DESC
            LIMIT 50
        """, (user_login,))
        history = cursor.fetchall()
        cursor.close()
        conn.close()
        return history
    except Error as e:
        print(f"Ошибка получения истории: {e}")
        return []


def clear_user_history(user_login):
    """Очищает историю поиска пользователя"""
    conn = get_history_connection()
    if not conn:
        return False

    try:
        cursor = conn.cursor()
        cursor.execute("""
            DELETE sr FROM search_results sr
            INNER JOIN search_queries sq ON sr.query_id = sq.id
            WHERE sq.user_login = %s
        """, (user_login,))
        cursor.execute("DELETE FROM search_queries WHERE user_login = %s", (user_login,))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Error as e:
        print(f"Ошибка очистки истории: {e}")
        return False


# ==================== ГЛАВНОЕ ПРИЛОЖЕНИЕ ====================
class AuthFrame(tk.Frame):
    def __init__(self, master, on_success, **kwargs):
        super().__init__(master, **kwargs)
        self.on_success = on_success
        self.pack(fill=tk.BOTH, expand=True)
        self.current_view = None
        self.message_label = None
        self.current_user = None
        master.geometry("500x500")
        master.resizable(False, False)
        self.create_select_view()

    def clear_view(self):
        for widget in self.winfo_children():
            widget.destroy()

    def show_message(self, text, color="black"):
        self.clear_message()
        self.message_label = tk.Label(
            self,
            text=text,
            font=("Arial", 11),
            bg="white",
            fg=color,
            pady=10,
            wraplength=450
        )
        self.message_label.pack(pady=(0, 15))

    def clear_message(self):
        if self.message_label:
            self.message_label.destroy()
            self.message_label = None

    def create_select_view(self):
        self.clear_view()
        self.clear_message()
        self.configure(bg="white")

        title = tk.Label(
            self,
            text="Добро пожаловать!",
            font=("Arial", 22, "bold"),
            bg="white"
        )
        title.pack(pady=(30, 30))

        container = tk.Frame(self, bg="white")
        container.pack(expand=True, pady=(0, 50))

        login_btn = tk.Button(
            container,
            text="Вход",
            font=("Arial", 18, "bold"),
            bg="#003366",
            fg="white",
            activebackground="#004c99",
            activeforeground="white",
            relief=tk.SOLID,
            borderwidth=2,
            width=16,
            height=1,
            command=self.create_login_view
        )
        login_btn.pack(pady=(10, 25), ipady=8)

        register_btn = tk.Button(
            container,
            text="Регистрация",
            font=("Arial", 18, "bold"),
            bg="#003366",
            fg="white",
            activebackground="#004c99",
            activeforeground="white",
            relief=tk.SOLID,
            borderwidth=2,
            width=16,
            height=1,
            command=self.create_register_view
        )
        register_btn.pack(pady=(25, 10), ipady=8)

        exit_btn = tk.Button(
            self,
            text="Выход",
            font=("Arial", 12),
            bg="#AAAAAA",
            fg="#666666",
            activebackground="#bbbbbb",
            activeforeground="#666666",
            relief=tk.FLAT,
            width=7,
            height=1,
            command=self.master.quit
        )
        exit_btn.place(x=10, rely=1.0, y=-25, anchor="sw")

        self.current_view = "select"

    def create_login_view(self):
        self.clear_view()
        self.clear_message()
        self.configure(bg="white")

        label = tk.Label(self, text="Вход", font=("Arial", 20, "bold"), bg="white")
        label.pack(pady=(30, 20))

        tk.Label(self, text="Логин:", bg="white", font=("Arial", 12)).pack(pady=5)
        self.login_entry = tk.Entry(self, font=("Arial", 12), width=25)
        self.login_entry.pack(pady=5)
        self.login_entry.focus()

        tk.Label(self, text="Пароль:", bg="white", font=("Arial", 12)).pack(pady=(15, 5))
        self.password_entry = tk.Entry(self, show="*", font=("Arial", 12), width=25)
        self.password_entry.pack(pady=5)
        self.password_entry.bind("<Return>", lambda e: self.login_action())

        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=25)

        back_btn = tk.Button(btn_frame, text="Назад", width=12, font=("Arial", 10), command=self.create_select_view)
        back_btn.pack(side=tk.LEFT, padx=8)

        login_btn = tk.Button(
            btn_frame, text="Войти", width=12, font=("Arial", 10, "bold"),
            bg="#003366", fg="white", command=self.login_action
        )
        login_btn.pack(side=tk.LEFT, padx=8)

        exit_btn = tk.Button(
            self, text="Выход", font=("Arial", 12), bg="#AAAAAA", fg="#666666",
            activebackground="#bbbbbb", relief=tk.FLAT, width=7, height=1,
            command=self.master.quit
        )
        exit_btn.place(x=10, rely=1.0, y=-25, anchor="sw")
        self.current_view = "login"

    def create_register_view(self):
        self.clear_view()
        self.clear_message()
        self.configure(bg="white")

        label = tk.Label(self, text="Регистрация", font=("Arial", 20, "bold"), bg="white")
        label.pack(pady=(25, 20))

        tk.Label(self, text="Логин (мин. 3 символа):", bg="white", font=("Arial", 12)).pack(pady=5)
        self.reg_login_entry = tk.Entry(self, font=("Arial", 12), width=25)
        self.reg_login_entry.pack(pady=5)
        self.reg_login_entry.focus()

        tk.Label(self, text="Пароль (мин. 4 символа):", bg="white", font=("Arial", 12)).pack(pady=(15, 5))
        self.reg_password_entry = tk.Entry(self, show="*", font=("Arial", 12), width=25)
        self.reg_password_entry.pack(pady=5)

        tk.Label(self, text="Подтверждение пароля:", bg="white", font=("Arial", 12)).pack(pady=(15, 5))
        self.reg_password_confirm_entry = tk.Entry(self, show="*", font=("Arial", 12), width=25)
        self.reg_password_confirm_entry.pack(pady=5)
        self.reg_password_confirm_entry.bind("<Return>", lambda e: self.register_action())

        btn_frame = tk.Frame(self, bg="white")
        btn_frame.pack(pady=25)

        back_btn = tk.Button(btn_frame, text="Назад", width=12, font=("Arial", 10), command=self.create_select_view)
        back_btn.pack(side=tk.LEFT, padx=8)

        register_btn = tk.Button(
            btn_frame, text="Зарегистрироваться", width=18, font=("Arial", 10, "bold"),
            bg="#003366", fg="white", command=self.register_action
        )
        register_btn.pack(side=tk.LEFT, padx=8)

        exit_btn = tk.Button(
            self, text="Выход", font=("Arial", 12), bg="#AAAAAA", fg="#666666",
            activebackground="#bbbbbb", relief=tk.FLAT, width=7, height=1,
            command=self.master.quit
        )
        exit_btn.place(x=10, rely=1.0, y=-25, anchor="sw")
        self.current_view = "register"

    def login_action(self):
        login = self.login_entry.get().strip()
        password = self.password_entry.get()

        if not login or not password:
            self.show_message("Введите логин и пароль!", "red")
            self.login_entry.focus()
            return

        if check_login(login, password):
            self.current_user = login
            self.show_message(f"Добро пожаловать, {login}!", "green")
            self.master.after(1500, self.on_success)
        else:
            self.show_message("Неверный логин или пароль!", "red")
            self.password_entry.delete(0, tk.END)
            self.login_entry.select_range(0, tk.END)
            self.login_entry.focus()

    def register_action(self):
        login = self.reg_login_entry.get().strip()
        pwd = self.reg_password_entry.get()
        pwd_confirm = self.reg_password_confirm_entry.get()

        self.clear_message()

        if not login or not pwd or not pwd_confirm:
            self.show_message("Заполните все поля!", "red")
            return

        if len(login) < 3:
            self.show_message("Логин должен быть не короче 3 символов!", "red")
            self.reg_login_entry.focus()
            return

        if len(pwd) < 4:
            self.show_message("Пароль должен быть не короче 4 символов!", "red")
            self.reg_password_entry.focus()
            return

        if pwd != pwd_confirm:
            self.show_message("Пароли не совпадают!", "red")
            self.reg_password_entry.delete(0, tk.END)
            self.reg_password_confirm_entry.delete(0, tk.END)
            self.reg_password_entry.focus()
            return

        if register_user(login, pwd):
            self.show_message(f"'{login}' зарегистрирован!\nПереходим ко входу...", "green")
            self.master.after(1500, lambda: self._after_register(login))
        else:
            self.show_message("Пользователь с таким логином уже существует!", "red")
            self.reg_login_entry.delete(0, tk.END)
            self.reg_login_entry.focus()

    def _after_register(self, login):
        self.create_login_view()
        self.login_entry.delete(0, tk.END)
        self.login_entry.insert(0, login)
        self.login_entry.focus()
        self.clear_message()


class YandexSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Программа интеллектуального поиска в интернете")
        self.root.configure(bg="white")
        self.links = {}
        self.results_data = []
        self.history_mode = False
        self.current_user = None

        # Проверяем подключение к базам данных
        self.check_databases()

        self.auth_frame = AuthFrame(root, self.after_auth)

    def check_databases(self):
        """Проверяет подключение к базам данных"""
        users_conn = get_users_connection()
        if users_conn:
            print("Подключение к базе users_db успешно")
            users_conn.close()
        else:
            messagebox.showerror("Ошибка",
                                 "Не удалось подключиться к базе данных users_db!\nПроверьте настройки MySQL.")
            self.root.quit()

        history_conn = get_history_connection()
        if history_conn:
            print("Подключение к базе search_history_db успешно")
            history_conn.close()
        else:
            messagebox.showerror("Ошибка",
                                 "Не удалось подключиться к базе данных search_history_db!\nПроверьте настройки MySQL.")
            self.root.quit()

    def after_auth(self):
        self.current_user = self.auth_frame.current_user
        self.auth_frame.destroy()
        self.root.geometry("950x720")
        self.root.resizable(True, True)
        self.build_main_interface()

    def build_main_interface(self):
        self.root.configure(bg="white")
        self.links = {}
        self.results_data = []
        self.history_mode = False

        self.title_label = tk.Label(
            self.root,
            text="Программа интеллектуального поиска в интернете",
            font=("Arial", 18, "bold"),
            bg="white",
            wraplength=900,
            justify="center"
        )
        self.title_label.pack(pady=(20, 15))

        self.user_info = tk.Label(
            self.root,
            text=f"Пользователь: {self.current_user}",
            font=("Arial", 10),
            bg="white",
            fg="#666666"
        )
        self.user_info.pack(pady=(0, 5))

        self.input_frame = tk.Frame(self.root, bg="white")
        self.input_frame.pack(fill=tk.X, padx=30, pady=(0, 30))
        self.input_frame.columnconfigure(0, weight=1)
        self.input_frame.columnconfigure(1, weight=0)

        self.search_entry = tk.Entry(
            self.input_frame,
            font=("Arial", 16),
            bg="#f3f3f3",
            relief=tk.GROOVE,
            borderwidth=3
        )
        self.search_entry.grid(row=0, column=0, sticky="we", padx=(0, 10), ipady=10)
        self.search_entry.bind("<Return>", lambda event: self.start_search())

        self.search_button = tk.Button(
            self.input_frame,
            text="Поиск",
            font=("Arial", 18, "bold"),
            bg="#0054b9",
            fg="white",
            activebackground="#196be4",
            activeforeground="white",
            relief=tk.RAISED,
            borderwidth=0,
            command=self.start_search,
            height=1,
            width=9
        )
        self.search_button.grid(row=0, column=1, sticky="e")

        self.result_text = scrolledtext.ScrolledText(
            self.root,
            wrap=tk.WORD,
            font=("Arial", 12),
            bg="#f9f9f9",
            borderwidth=1,
            relief=tk.FLAT,
            height=25
        )
        self.result_text.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 10))
        self.result_text.tag_configure("title", font=("Arial", 12, "bold"))
        self.result_text.tag_configure("link", foreground="#0054b9", underline=1)
        self.result_text.tag_configure("description", font=("Arial", 11))
        self.result_text.tag_configure("debug", foreground="gray")
        self.result_text.tag_configure("history_item", font=("Arial", 11))

        self.export_frame = tk.Frame(self.root, bg="white")
        self.export_frame.pack(fill=tk.X, padx=30, pady=(0, 10))

        self.history_button = tk.Button(
            self.export_frame,
            text="История запросов",
            font=("Arial", 11, "bold"),
            bg="#17a2b8",
            fg="white",
            activebackground="#138496",
            relief=tk.RAISED,
            borderwidth=0,
            command=self.show_history,
            width=18
        )
        self.history_button.pack(side=tk.LEFT, padx=(0, 10), pady=8)

        self.clear_history_button = tk.Button(
            self.export_frame,
            text="Очистить историю",
            font=("Arial", 11, "bold"),
            bg="#dc3545",
            fg="white",
            activebackground="#c82333",
            relief=tk.RAISED,
            borderwidth=0,
            command=self.clear_history,
            width=18
        )

        self.back_button = tk.Button(
            self.export_frame,
            text="Назад к поиску",
            font=("Arial", 11, "bold"),
            bg="#6c757d",
            fg="white",
            activebackground="#5a6268",
            relief=tk.RAISED,
            borderwidth=0,
            command=self.back_to_search,
            width=15
        )

        self.export_button = tk.Button(
            self.export_frame,
            text="Выгрузить в Excel",
            font=("Arial", 11, "bold"),
            bg="#28a745",
            fg="white",
            activebackground="#218838",
            relief=tk.RAISED,
            borderwidth=0,
            command=self.export_to_excel,
            width=20
        )
        self.export_button.pack(side=tk.RIGHT, pady=8)

        self.status_var = tk.StringVar()
        self.status_var.set("Готов к поиску")
        self.status_label = tk.Label(
            self.root,
            textvariable=self.status_var,
            font=("Arial", 11),
            bg="#e9ecef",
            fg="#495057",
            anchor=tk.W,
            padx=15,
            pady=8
        )
        self.status_label.pack(fill=tk.X, pady=(0, 10))

    def show_history(self):
        """Показывает историю запросов пользователя"""
        try:
            self.history_mode = True

            self.history_button.pack_forget()
            self.clear_history_button.pack(side=tk.LEFT, padx=(0, 10), pady=8)
            self.back_button.pack(side=tk.LEFT, padx=(0, 10), pady=8)

            self.result_text.config(state="normal")
            self.result_text.delete(1.0, tk.END)

            history = get_user_history(self.current_user)

            if not history:
                self.result_text.insert(tk.END, "История запросов пуста.", "history_item")
                self.status_var.set("История запросов пуста")
            else:
                # Выводим историю без заголовков и тире
                for row in history:
                    query_id, query, date, count = row
                    try:
                        date_str = date.strftime("%d.%m.%Y %H:%M:%S") if hasattr(date, 'strftime') else str(date)
                    except:
                        date_str = str(date)

                    tag_name = f"history_query_{query_id}"

                    self.result_text.insert(tk.END, f"[{date_str}] ", "history_item")

                    start_idx = self.result_text.index(tk.INSERT)
                    self.result_text.insert(tk.END, f"{query}", tag_name)
                    end_idx = self.result_text.index(tk.INSERT)

                    self.result_text.tag_add(tag_name, start_idx, end_idx)
                    self.result_text.tag_configure(tag_name, foreground="#0054b9", underline=1)
                    self.result_text.tag_bind(tag_name, "<Button-1>", lambda e, q=query: self.repeat_search(q))

                    self.result_text.insert(tk.END, f" ({count} результатов)\n\n", "history_item")

            self.result_text.config(state="disabled")
            self.status_var.set("Режим просмотра истории")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить историю: {e}")
            self.back_to_search()

    def repeat_search(self, query):
        self.back_to_search()
        self.search_entry.delete(0, tk.END)
        self.search_entry.insert(0, query)
        self.start_search()

    def clear_history(self):
        try:
            result = messagebox.askyesno("Подтверждение", "Вы уверены, что хотите очистить всю историю запросов?")
            if result:
                if clear_user_history(self.current_user):
                    self.status_var.set("История очищена")
                    if self.history_mode:
                        self.show_history()
                    else:
                        messagebox.showinfo("Готово", "История запросов очищена")
                else:
                    messagebox.showerror("Ошибка", "Не удалось очистить историю")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось очистить историю: {e}")

    def back_to_search(self):
        self.history_mode = False

        self.clear_history_button.pack_forget()
        self.back_button.pack_forget()
        self.history_button.pack(side=tk.LEFT, padx=(0, 10), pady=8)

        self.result_text.config(state="normal")
        self.result_text.delete(1.0, tk.END)
        self.result_text.config(state="disabled")

        self.status_var.set("Готов к поиску")

    def start_search(self):
        if self.history_mode:
            self.back_to_search()
        self.search_button.config(state="disabled")
        threading.Thread(target=self.search, daemon=True).start()

    def get_available_browser(self):
        """Автоматически определяет доступный браузер"""
        browsers = [
            ('chrome', 'google-chrome', 'chrome', 'Chrome', ChromeDriverManager, ChromeOptions, ChromeService),
            ('firefox', 'firefox', 'firefox', 'Firefox', GeckoDriverManager, FirefoxOptions, FirefoxService),
            ('edge', 'edge', 'edge', 'Edge', EdgeChromiumDriverManager, EdgeOptions, EdgeService),
        ]

        # Поиск установленных браузеров
        for browser_name, linux_name, mac_name, display_name, driver_manager, options_class, service_class in browsers:
            try:
                # Проверяем наличие браузера
                if sys.platform == 'win32':
                    if browser_name == 'chrome':
                        paths = [
                            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                            os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
                        ]
                        for path in paths:
                            if os.path.exists(path):
                                return (display_name, driver_manager, options_class, service_class)
                    elif browser_name == 'firefox':
                        paths = [
                            r"C:\Program Files\Mozilla Firefox\firefox.exe",
                            r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
                            os.path.expanduser(r"~\AppData\Local\Mozilla Firefox\firefox.exe")
                        ]
                        for path in paths:
                            if os.path.exists(path):
                                return (display_name, driver_manager, options_class, service_class)
                    elif browser_name == 'edge':
                        paths = [
                            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
                        ]
                        for path in paths:
                            if os.path.exists(path):
                                return (display_name, driver_manager, options_class, service_class)
                else:
                    # Для Linux/Mac
                    result = subprocess.run(['which', linux_name], capture_output=True, text=True)
                    if result.returncode == 0:
                        return (display_name, driver_manager, options_class, service_class)
            except:
                continue

        return None

    def extract_real_url(self, url):
        """Извлекает реальный URL из ссылки Яндекса"""
        if not url:
            return None

        # Если это редирект Яндекса
        if url.startswith('/url?q=') or 'yandex.ru/url?q=' in url:
            try:
                parsed = urlparse(url)
                params = parse_qs(parsed.query)
                if 'q' in params:
                    return params['q'][0]
            except:
                pass

        # Если относительная ссылка
        if url.startswith('/'):
            return "https://yandex.ru" + url

        return url

    def search(self):
        query = self.search_entry.get().strip()
        if not query:
            self.status_var.set("Введите поисковый запрос")
            self.search_button.config(state="normal")
            return

        self.status_var.set(f"Поиск: {query}...")
        self.result_text.config(state="normal")
        self.result_text.delete(1.0, tk.END)
        self.results_data = []
        self.root.update()

        query_id = insert_search_query(self.current_user, query)

        driver = None
        try:
            # Автоматически определяем доступный браузер
            browser_info = self.get_available_browser()

            if not browser_info:
                self.result_text.insert(tk.END,
                                        "Не найден ни один поддерживаемый браузер!\nУстановите Chrome, Firefox или Edge\n",
                                        "debug")
                self.status_var.set("Ошибка: браузер не найден")
                return

            browser_name, driver_manager, options_class, service_class = browser_info
            self.status_var.set(f"Используется браузер: {browser_name}")
            print(f"Используется браузер: {browser_name}")

            # Настройки для браузера
            options = options_class()
            options.add_argument("--headless")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")

            # User-Agent для имитации реального пользователя
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            if browser_name == 'Chrome':
                user_agent += " Chrome/120.0.0.0 Safari/537.36"
            elif browser_name == 'Firefox':
                user_agent += " Firefox/121.0"
            elif browser_name == 'Edge':
                user_agent += " Edg/120.0.0.0"

            options.add_argument(f"--user-agent={user_agent}")

            # Устанавливаем драйвер
            service = service_class(driver_manager().install())

            # Запускаем браузер
            if browser_name == 'Firefox':
                driver = webdriver.Firefox(service=service, options=options)
            elif browser_name == 'Edge':
                driver = webdriver.Edge(service=service, options=options)
            else:
                driver = webdriver.Chrome(service=service, options=options)

            print(f"Браузер {browser_name} успешно запущен")

            # Поиск в Яндекс
            search_url = f"https://yandex.ru/search/?text={quote_plus(query)}&lr=213"
            driver.get(search_url)

            # Проверка на капчу
            if "captcha" in driver.current_url or "showcaptcha" in driver.page_source.lower():
                self.result_text.insert(tk.END, "Яндекс запросил капчу, возможна блокировка\n", "debug")
                self.status_var.set("Блокировка Яндекса")
                return

            # Ждем загрузки страницы
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".serp-item, .organic, .main__content"))
                )
            except Exception:
                pass

            time.sleep(2)

            # Получаем HTML
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # Множество селекторов для поиска результатов
            results_selectors = [
                'li.serp-item',
                'div.serp-item',
                '.organic',
                '.card',
                '.extended-text',
                '.result',
                '.search-result'
            ]

            search_results = []
            for selector in results_selectors:
                results = soup.select(selector)
                if results:
                    search_results.extend(results)

            # Удаляем дубликаты
            seen = set()
            unique_results = []
            for r in search_results:
                r_str = str(r)[:200]
                if r_str not in seen:
                    seen.add(r_str)
                    unique_results.append(r)

            print(f"Найдено уникальных результатов: {len(unique_results)}")

            found = 0
            for result in unique_results:
                if found >= 15:
                    break

                # Поиск заголовка
                title = None
                title_elem = None

                title_selectors = [
                    'h2 a',
                    'h3 a',
                    '.organic__title a',
                    '.organic__title-wrapper a',
                    '.OrganicTitle a',
                    '.title a',
                    'a.Link',
                    '.serp-item__title a',
                    '.serp-item__title-link',
                    '.result__title a',
                    '.search-result__title a',
                    '.card-title a'
                ]

                for selector in title_selectors:
                    title_elem = result.select_one(selector)
                    if title_elem:
                        title = title_elem.get_text().strip()
                        break

                if not title or len(title) < 5:
                    continue

                # Поиск ссылки
                link = ""
                if title_elem and title_elem.get('href'):
                    link = title_elem.get('href')
                else:
                    link_elem = result.select_one('a')
                    if link_elem and link_elem.get('href'):
                        link = link_elem.get('href')

                if not link:
                    continue

                # Извлекаем реальный URL
                link = self.extract_real_url(link)
                if not link or len(link) < 10:
                    continue

                # Поиск описания
                desc = "Описание отсутствует"
                desc_selectors = [
                    '.organic__text',
                    '.organic__content',
                    '.serp-item__text',
                    '.result__description',
                    '.search-result__snippet',
                    '.extended-text__content',
                    '.card-description',
                    '.text-container',
                    '.OrganicText'
                ]

                for selector in desc_selectors:
                    desc_elem = result.select_one(selector)
                    if desc_elem and desc_elem.get_text().strip():
                        desc = desc_elem.get_text().strip()
                        if len(desc) > 30:
                            break

                # Очистка текста
                title = ' '.join(title.split())
                desc = ' '.join(desc.split())
                desc = desc[:300] + "..." if len(desc) > 300 else desc

                # Сохраняем результат
                link_tag = f"link_{found + 1}"
                self.links[link_tag] = link

                self.results_data.append({
                    "№": found + 1,
                    "Заголовок": title,
                    "Описание": desc,
                    "Ссылка": link
                })

                # Выводим в интерфейс
                self.result_text.insert(tk.END, f"{found + 1}. ", "title")
                self.result_text.insert(tk.END, f"{title}\n", "title")
                self.result_text.insert(tk.END, f"{desc}\n\n", "description")

                start_idx = self.result_text.index(tk.INSERT)
                self.result_text.insert(tk.END, f"{link}\n\n", link_tag)
                end_idx = self.result_text.index(tk.INSERT)
                self.result_text.tag_add(link_tag, start_idx, end_idx)
                self.result_text.tag_configure(link_tag, foreground="#0054b9", underline=1)
                self.result_text.tag_bind(link_tag, "<Button-1>", self.open_link)

                found += 1
                self.root.update()
                print(f"Найден результат {found}: {title[:50]}...")

            if found == 0:
                self.result_text.insert(tk.END, "Результаты не найдены.\nПопробуйте другой запрос.\n", "debug")
                self.status_var.set("Результаты не найдены")
            else:
                self.status_var.set(f"Найдено {found} результатов | Сохранено в БД")

            if query_id and self.results_data:
                insert_search_results(query_id, self.results_data)

        except Exception as e:
            error_msg = f"Ошибка поиска: {str(e)}\n\n{traceback.format_exc()}"
            self.result_text.insert(tk.END, error_msg, "debug")
            self.status_var.set("Произошла ошибка")
            print(f"Подробная ошибка: {traceback.format_exc()}")

        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            self.result_text.config(state="normal")
            self.search_button.config(state="normal")

    def open_link(self, event):
        try:
            tag_name = event.widget.tag_names(tk.CURRENT)[0]
            url = self.links.get(tag_name)
            if url:
                webbrowser.open_new(url)
                self.status_var.set(f"Открыта ссылка: {url[:60]}...")
        except:
            pass

    def export_to_excel(self):
        if not self.results_data:
            messagebox.showwarning("Внимание", "Сначала выполните поиск!")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
            initialfile=f"search_results_{time.strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Результаты поиска"

                ws.merge_cells("A1:D1")
                ws["A1"] = f"Поиск: {self.search_entry.get() or 'Результаты'}"
                ws["A1"].font = Font(size=16, bold=True)
                ws["A1"].alignment = Alignment(horizontal="center")

                ws.cell(row=2, column=1, value=f"Пользователь: {self.current_user}")
                ws.cell(row=2, column=1).font = Font(size=10, italic=True)
                ws.cell(row=3, column=1, value=f"Дата: {time.strftime('%d.%m.%Y %H:%M:%S')}")
                ws.cell(row=3, column=1).font = Font(size=10, italic=True)

                headers = ["№", "Заголовок", "Описание", "Ссылка"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="003366")
                    cell.alignment = Alignment(horizontal="center")

                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 60
                ws.column_dimensions['D'].width = 50

                for row, item in enumerate(self.results_data, 6):
                    ws[f"A{row}"].value = item["№"]
                    ws[f"B{row}"].value = item["Заголовок"]
                    ws[f"C{row}"].value = item["Описание"]
                    cell_d = ws[f"D{row}"]
                    cell_d.value = item["Ссылка"]
                    cell_d.hyperlink = item["Ссылка"]
                    cell_d.style = "Hyperlink"

                wb.save(filename)
                messagebox.showinfo("Готово!", f"Сохранено: {filename}")
                self.status_var.set(f"Excel сохранен: {filename.split('/')[-1]}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = YandexSearchApp(root)
    root.mainloop()